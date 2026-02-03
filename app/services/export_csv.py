"""
CSV Export Service
Exports processed products to Shopify and Business Central CSV formats
Works with products from ANY source: CSV, PDF, URL, image, text, barcode
"""
import csv
import io
import logging
import re
from typing import List, Dict, Any
import pandas as pd

from .shopify_mapper import map_to_shopify_csv, SHOPIFY_CSV_HEADERS
from .csv_parser import clean_barcode

logger = logging.getLogger(__name__)


def clean_text(text: str) -> str:
    """Remove artifacts like <!--image--> from text"""
    if not text:
        return ""
    # Remove HTML comments
    text = re.sub(r'<!--[^>]*-->', '', text)
    # Clean up extra whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def export_to_business_central(products: List[Dict[str, Any]]) -> bytes:
    """
    Export products to Business Central CSV format using pandas
    
    Format:
    - UTF-8 BOM encoding
    - HTML formatted descriptions (<p>, <br> tags)
    - NO image columns (removed per business requirement)
    - 6 columns: SKU, Barcode, Description, Net Weight (KG), Short Description, Long Description
    
    Args:
        products: List of product dictionaries from ANY source
    
    Returns:
        CSV file content as bytes with UTF-8 BOM
    """
    
    rows = []
    
    for product in products:
        descriptions = product.get('descriptions', {})
        specs = product.get('specifications', {})
        
        # Get and clean descriptions
        short_desc = clean_text(descriptions.get('shortDescription', ''))
        long_desc = clean_text(descriptions.get('longDescription', ''))
        product_name = clean_text(product.get('name', ''))
        
        # Format short description as HTML bullet points
        short_desc_html = format_short_description_html(short_desc)
        
        # Format long description as HTML paragraphs
        long_desc_html = format_long_description_html(long_desc)
        
        sku = product.get('sku', product.get('id', 'unknown'))
        
        # Get weight (try multiple sources)
        weight = (
            specs.get('weight', '') or
            specs.get('weightKg', '') or
            specs.get('Net Weight (KG)', '') or
            ''
        )
        
        # Strip "kg" suffix if present
        if isinstance(weight, str):
            weight = weight.replace('kg', '').replace('KG', '').strip()
        
        # Clean barcode to prevent scientific notation in export
        raw_barcode = product.get('barcode', '')
        cleaned_barcode = clean_barcode(raw_barcode) if raw_barcode else ''

        # Build row - NO IMAGE COLUMNS
        row = {
            'SKU': sku,
            'Barcode': cleaned_barcode,
            'Description': product_name,
            'Net Weight (KG)': weight,
            'Short Description': short_desc_html,
            'Long Description': long_desc_html,
        }
        
        rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Ensure Barcode column is string type to prevent scientific notation
    if 'Barcode' in df.columns:
        df['Barcode'] = df['Barcode'].astype(str)

    # Export to CSV with UTF-8 BOM
    csv_buffer = io.StringIO()
    df.to_csv(
        csv_buffer,
        index=False,
        encoding='utf-8-sig',
        lineterminator='\n',
        quoting=csv.QUOTE_MINIMAL,
    )
    
    csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')
    
    logger.info(f"Exported {len(products)} products to Business Central CSV format")

    return csv_bytes


def export_to_shopify(products: List[Dict[str, Any]]) -> bytes:
    """
    Export products to Shopify CSV format with metafields

    Format:
    - UTF-8 BOM encoding
    - Shopify-compatible column headers with metafield notation
    - Dietary preferences, allergens, ingredients as metafields

    Columns:
    - ID, Handle, Title, Body HTML, Vendor, Type, Variant Barcode
    - Metafield: custom.allergens [list.single_line_text_field]
    - Metafield: pdp.ingredients [rich_text_field]
    - Metafield: pdp.nutrition [list.single_line_text_field]
    - Metafield: custom.dietary_preferences [list.single_line_text_field]
    - Metafield: custom.brand [single_line_text_field]

    Args:
        products: List of product dictionaries from ANY source

    Returns:
        CSV file content as bytes with UTF-8 BOM
    """

    rows = []

    for product in products:
        # Map product to Shopify format
        shopify_row = map_to_shopify_csv(product)
        rows.append(shopify_row)

    # Create DataFrame with ordered columns
    df = pd.DataFrame(rows, columns=SHOPIFY_CSV_HEADERS)

    # Ensure barcode column is string type to prevent scientific notation
    if 'Variant Barcode' in df.columns:
        df['Variant Barcode'] = df['Variant Barcode'].astype(str)

    # Export to CSV with UTF-8 BOM
    csv_buffer = io.StringIO()
    df.to_csv(
        csv_buffer,
        index=False,
        encoding='utf-8-sig',
        lineterminator='\n',
        quoting=csv.QUOTE_MINIMAL,
    )

    csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')

    logger.info(f"Exported {len(products)} products to Shopify CSV format")

    return csv_bytes


def format_short_description_html(text: str) -> str:
    """
    Format short description as HTML bullet points
    
    Input: "Feature 1\nFeature 2\nFeature 3"
    Output: "<p>Feature 1<br>Feature 2<br>Feature 3</p>"
    """
    if not text:
        return ""
    
    # If already HTML formatted, return as-is
    if text.strip().startswith('<p>'):
        return text
    
    # Split into lines
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    if not lines:
        return ""
    
    # Join with line breaks (no trailing commas)
    html = '<p>' + '<br>'.join(lines) + '</p>'
    
    return html


def format_long_description_html(text: str) -> str:
    """
    Format long description as HTML paragraphs
    
    Input: "Paragraph 1\n\nParagraph 2"
    Output: "<p>Paragraph 1</p><p>Paragraph 2</p>"
    """
    if not text:
        return ""
    
    # If already HTML formatted, return as-is
    if '<p>' in text:
        return text
    
    # Split into paragraphs (double newlines)
    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
    
    if not paragraphs:
        return f'<p>{text.strip()}</p>'
    
    html = ''.join([f'<p>{p}</p>' for p in paragraphs])
    
    return html


def sanitize_filename(name: str) -> str:
    """Sanitize product name for use in filename"""
    name = name.replace(' ', '-')
    name = re.sub(r'[^\w\-]', '', name)
    name = name[:50]
    return name


def export_to_excel(products: List[Dict[str, Any]]) -> bytes:
    """
    Export products to Excel format with professional formatting

    Features:
    - Bold headers with background color
    - Auto-width columns (capped for readability)
    - Text wrapping for description columns
    - Frozen header row
    - No image columns
    - Includes dietary preferences, ingredients, allergens
    """

    rows = []

    for product in products:
        descriptions = product.get('descriptions', {})
        specs = product.get('specifications', {})

        # Get plain text versions (strip HTML for Excel)
        short_desc = strip_html(clean_text(descriptions.get('shortDescription', '')))
        long_desc = strip_html(clean_text(descriptions.get('longDescription', '')))
        meta_desc = strip_html(clean_text(descriptions.get('metaDescription', '')))
        product_name = clean_text(product.get('name', ''))

        features = product.get('features', [])

        # Get weight
        weight = (
            specs.get('weight', '') or
            specs.get('weightKg', '') or
            ''
        )
        if isinstance(weight, str):
            weight = weight.replace('kg', '').replace('KG', '').strip()

        # Get dietary preferences (from CSV extraction or enrichment)
        dietary = product.get('dietary') or product.get('dietary_preferences') or descriptions.get('dietary_preferences', [])
        dietary_str = ', '.join(dietary) if isinstance(dietary, list) else str(dietary) if dietary else ''

        # Get ingredients
        ingredients = product.get('ingredients', '')
        if isinstance(ingredients, list):
            ingredients = ', '.join(ingredients)

        # Get allergens
        allergens = product.get('allergens', [])
        allergens_str = ', '.join(allergens) if isinstance(allergens, list) else str(allergens) if allergens else ''

        # Get Earthfare icons (Palm Oil Free, Organic, Vegan, Fairtrade)
        icons = descriptions.get('icons', []) or product.get('icons', [])
        icons_str = ', '.join(icons) if isinstance(icons, list) else str(icons) if icons else ''

        # Get nutrition data (per 100g)
        nutrition = product.get('nutrition', {})
        nutrition_shopify = product.get('nutrition_shopify', [])
        nutrition_source = product.get('nutrition_source', '')

        # Format nutrition for display
        if nutrition_shopify and isinstance(nutrition_shopify, list):
            nutrition_str = '\n'.join(nutrition_shopify)
        elif nutrition and isinstance(nutrition, dict):
            # Format from dict
            nutrition_lines = []
            nutrition_order = [
                ('energy_kcal', 'Energy', 'kcal'),
                ('fat', 'Fat', 'g'),
                ('saturates', 'Saturates', 'g'),
                ('carbohydrates', 'Carbohydrates', 'g'),
                ('sugars', 'Sugars', 'g'),
                ('fibre', 'Fibre', 'g'),
                ('protein', 'Protein', 'g'),
                ('salt', 'Salt', 'g'),
            ]
            for key, label, unit in nutrition_order:
                if key in nutrition and nutrition[key]:
                    nutrition_lines.append(f"{label}: {nutrition[key]}{unit}")
            nutrition_str = '\n'.join(nutrition_lines)
        else:
            nutrition_str = ''

        # Clean barcode to prevent scientific notation in export
        raw_barcode = product.get('barcode', '')
        cleaned_barcode = clean_barcode(raw_barcode) if raw_barcode else ''

        row = {
            'SKU': product.get('sku', ''),
            'Barcode': cleaned_barcode,
            'Product Name': product_name,
            'Brand': product.get('brand', ''),
            'Category': product.get('category', ''),
            'Weight (KG)': weight,
            'Short Description': short_desc,
            'Long Description': long_desc,
            'Meta Description': meta_desc,
            'Dietary Preferences': dietary_str,
            'Icons': icons_str,
            'Ingredients': ingredients,
            'Allergens': allergens_str,
            'Nutrition (per 100g)': nutrition_str,
            'Nutrition Source': nutrition_source,
            'Features': '\n'.join(features) if features else '',
        }

        rows.append(row)

    df = pd.DataFrame(rows)

    # Ensure Barcode column is string type to prevent scientific notation
    if 'Barcode' in df.columns:
        df['Barcode'] = df['Barcode'].astype(str)

    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Products')
        
        worksheet = writer.sheets['Products']
        
        # Import styles
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Column width settings (column letter: width)
        column_widths = {
            'A': 15,   # SKU
            'B': 18,   # Barcode
            'C': 40,   # Product Name
            'D': 15,   # Brand
            'E': 20,   # Category
            'F': 12,   # Weight
            'G': 50,   # Short Description
            'H': 70,   # Long Description
            'I': 50,   # Meta Description
            'J': 30,   # Dietary Preferences
            'K': 35,   # Icons (Palm Oil Free, Organic, Vegan, Fairtrade)
            'L': 60,   # Ingredients
            'M': 30,   # Allergens
            'N': 40,   # Nutrition (per 100g)
            'O': 18,   # Nutrition Source
            'P': 40,   # Features
        }
        
        # Apply column widths
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # Set row heights for better readability
        worksheet.row_dimensions[1].height = 25  # Header row
        for row_num in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_num].height = 60  # Data rows
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    excel_buffer.seek(0)
    
    logger.info(f"Exported {len(products)} products to Excel format")
    
    return excel_buffer.read()


def strip_html(text: str) -> str:
    """Remove HTML tags from text"""
    if not text:
        return ""
    
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', ' ', text)
    
    # Decode HTML entities
    text = text.replace('&nbsp;', ' ')
    text = text.replace('&amp;', '&')
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    text = text.replace('&quot;', '"')
    text = text.replace('&#39;', "'")
    
    # Clean up whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text
